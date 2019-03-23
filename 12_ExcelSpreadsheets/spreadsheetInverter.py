import openpyxl
import sys

spreadSheet = sys.argv[1]

wb = openpyxl.load_workbook(spreadSheet)
wb2 = openpyxl.Workbook()

sheet = wb.active
sheet2 = wb2.active


for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        sheet2[openpyxl.utils.get_column_letter(row) + str(column)].value = sheet[openpyxl.utils.get_column_letter(column) + str(row)].value


wb2.save('inverted.xlsx')