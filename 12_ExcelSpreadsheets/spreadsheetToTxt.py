import openpyxl
import sys

spreadsheet = sys.argv[1]

wb = openpyxl.load_workbook(spreadsheet)
sheet = wb.active

counter = 1
for column in range(1, sheet.max_column + 1):
    fileName = "Column%s.txt" % (openpyxl.utils.get_column_letter(counter))
    f = open(fileName, 'w')
    for row in range(1, sheet.max_row + 1):
        print(sheet[openpyxl.utils.get_column_letter(column) + str(row)].value)
        f.write(str(sheet[openpyxl.utils.get_column_letter(column) + str(row)].value) + '\n')
    counter += 1
    f.close()