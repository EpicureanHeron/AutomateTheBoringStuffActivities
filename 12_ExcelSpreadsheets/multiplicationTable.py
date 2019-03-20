import openpyxl
from openpyxl.styles import Font
import sys

wb = openpyxl.Workbook()
sheet = wb.active
fontObj1 = Font(name='Times New Roman', bold=True)

for i in range(1, int(sys.argv[1]) + 1): 
    sheet['A' + str(i+1)] = i
    sheet['A' + str(i+1)].font = fontObj1
    sheet[openpyxl.utils.get_column_letter(i+1) + '1'] = i
    sheet[openpyxl.utils.get_column_letter(i+1) + '1'].font = fontObj1

for row in range(2, sheet.max_row + 1):
    for column in range(2, sheet.max_column + 1):

        sheet[openpyxl.utils.get_column_letter(column) + str(row)] = sheet['A' + str(row)].value * sheet[openpyxl.utils.get_column_letter(column) + '1'].value

wb.save('multiTables.xlsx')