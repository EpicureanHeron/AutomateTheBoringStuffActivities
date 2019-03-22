import openpyxl
import sys

rowNumber = int(sys.argv[1])
blanksToInsert = int(sys.argv[2])
spreadSheet = sys.argv[3]



wb = openpyxl.load_workbook(spreadSheet)
wb2 = openpyxl.Workbook()

sheet = wb.active
sheet2 = wb2.active

for row in range(1, rowNumber+1):
    for column in range(1, sheet.max_column + 1):
        sheet2[openpyxl.utils.get_column_letter(column) + str(row)].value = sheet[openpyxl.utils.get_column_letter(column) + str(row)].value

for row in range(rowNumber+1, rowNumber + 1 + blanksToInsert):
    for column in range(1, sheet.max_column + 1):
         sheet2[openpyxl.utils.get_column_letter(column) + str(row)].value = ''

for row in range(rowNumber + 1 , sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        sheet2[openpyxl.utils.get_column_letter(column) + str(row + blanksToInsert)].value = sheet[openpyxl.utils.get_column_letter(column) + str(row)].value       

wb2.save('copied.xlsx')
# capture and save data from spreadsheet 

# insert # of rows from blanksToInsert variable

# remaining data from spreadsheet

# the first portion + blanks + second portion to new spreadsheet

# save spreadsheet