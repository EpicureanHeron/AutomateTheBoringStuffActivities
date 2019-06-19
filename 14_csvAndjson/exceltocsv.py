#! python3

import openpyxl
import csv
import os
import sys

#os.makedirs('excelToCSV', exist_ok=True)

counter = 0
for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue
    
    print('converting  ' + excelFile)
    wb = openpyxl.load_workbook(excelFile)

    sheetNames = wb.get_sheet_names()

    for sheet in sheetNames:

        outputFile = open('%s-%s.csv' % (excelFile[:-5], sheet), 'w', newline='')
        outputWriter = csv.writer(outputFile)
        counter += 1

        for row in range(2, wb[sheet].max_row+ 1):
            csvRow = []
            for column in range(2, wb[sheet].max_column + 1):
                value = (wb[sheet][openpyxl.utils.get_column_letter(column) + str(row)].value) 
                csvRow.append(value)

            outputWriter.writerow(csvRow)
        
    outputFile.close()
        
    
print('created %s files' % (str(counter)))